# -*- coding: utf-8 -*-
"""
RED SI Pipeline Excel 导出
===========================

将清洗数据、模型输出、指标汇总、系数表、Dashboard 和情景汇总
导出为格式化的多 sheet Excel 文件。
"""

from __future__ import annotations

import pandas as pd


def export_excel(
    clean: pd.DataFrame,
    model_df: pd.DataFrame,
    summary_metrics: pd.DataFrame,
    coef_df: pd.DataFrame,
    dashboard_df: pd.DataFrame,
    scenario_summary_df: pd.DataFrame,
    scenario_dashboard_df: pd.DataFrame,
    output_xlsx: str,
) -> None:
    """导出多 sheet Excel 并应用基础格式化。"""
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        clean.to_excel(writer, sheet_name="01_Cleaned_Raw", index=False)
        summary_metrics.to_excel(writer, sheet_name="02_Model_Summary", index=False)
        coef_df.to_excel(writer, sheet_name="03_Coefficients", index=False)
        dashboard_df.to_excel(writer, sheet_name="04_Dashboard_Table", index=False)
        model_df.to_excel(writer, sheet_name="05_Model_Output", index=False)
        scenario_summary_df.to_excel(writer, sheet_name="06_Scenario_Summary", index=False)
        scenario_dashboard_df.to_excel(writer, sheet_name="07_Scenario_Dashboard", index=False)

    # 基础格式化
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(output_xlsx)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    forecast_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 12), 38)

    # 预测期行浅绿色：用 target 缺失判断
    for sheet_name in ["01_Cleaned_Raw", "05_Model_Output"]:
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        if "brand_search_index_mil" in headers:
            target_col = headers.index("brand_search_index_mil") + 1
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, target_col).value is None:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(r, c).fill = forecast_fill

    wb.active = wb.sheetnames.index("02_Model_Summary")
    wb.save(output_xlsx)
